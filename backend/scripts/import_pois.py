"""Import POI_Master rows from a read-only Excel workbook into PostgreSQL."""

from __future__ import annotations

import argparse
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import datetime
import json
from pathlib import Path
import sys
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.db.models import Poi  # noqa: E402
from app.db.session import SessionLocal  # noqa: E402

SOURCE_TIMEZONE = ZoneInfo("Asia/Shanghai")
REQUIRED_COLUMNS = {
    "poi_id",
    "poi_name",
    "alias",
    "longitude",
    "latitude",
    "amap_address",
    "amap_type",
    "source_note",
    "created_at",
    "updated_at",
}


@dataclass(frozen=True)
class PoiImportRow:
    poi_id: str
    poi_name: str
    alias: str | None
    address: str
    longitude: float
    latitude: float
    category: str
    source: str
    created_at: datetime
    updated_at: datetime


@dataclass(frozen=True)
class ImportResult:
    rows_read: int
    inserted: int
    updated: int


def _text(value: object, *, required: bool = False) -> str | None:
    text = str(value).strip() if value is not None else ""
    if required and not text:
        raise ValueError("Required POI text value is blank")
    return text or None


def _timestamp(value: object) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        parsed = datetime.fromisoformat(value.strip())
    else:
        raise ValueError(f"Invalid POI timestamp: {value!r}")
    return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=SOURCE_TIMEZONE)


def read_poi_rows(source_path: Path) -> list[PoiImportRow]:
    """Read POI_Master without ever writing back to the source workbook."""
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if "POI_Master" not in workbook.sheetnames:
            raise ValueError("Workbook does not contain POI_Master")
        sheet = workbook["POI_Master"]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values)]
        missing = REQUIRED_COLUMNS - set(headers)
        if missing:
            raise ValueError(f"POI_Master missing columns: {sorted(missing)}")
        indexes = {name: headers.index(name) for name in REQUIRED_COLUMNS}

        rows: list[PoiImportRow] = []
        seen_ids: set[str] = set()
        for row_number, row in enumerate(values, start=2):
            if not any(value not in (None, "") for value in row):
                continue
            poi_id = _text(row[indexes["poi_id"]], required=True)
            assert poi_id is not None
            if poi_id in seen_ids:
                raise ValueError(f"Duplicate poi_id at row {row_number}: {poi_id}")
            seen_ids.add(poi_id)

            longitude = float(row[indexes["longitude"]])
            latitude = float(row[indexes["latitude"]])
            if not -180 <= longitude <= 180 or not -90 <= latitude <= 90:
                raise ValueError(
                    f"Invalid coordinates at row {row_number}: {longitude}, {latitude}"
                )

            rows.append(
                PoiImportRow(
                    poi_id=poi_id,
                    poi_name=_text(row[indexes["poi_name"]], required=True) or "",
                    alias=_text(row[indexes["alias"]]),
                    address=_text(row[indexes["amap_address"]], required=True) or "",
                    longitude=longitude,
                    latitude=latitude,
                    category=_text(row[indexes["amap_type"]], required=True) or "",
                    source=_text(row[indexes["source_note"]], required=True) or "",
                    created_at=_timestamp(row[indexes["created_at"]]),
                    updated_at=_timestamp(row[indexes["updated_at"]]),
                )
            )
        return rows
    finally:
        workbook.close()


def upsert_pois(
    rows: Iterable[PoiImportRow],
    *,
    session_factory=SessionLocal,
) -> ImportResult:
    poi_rows = list(rows)
    with session_factory() as session:
        existing_ids = set(
            session.scalars(
                select(Poi.poi_id).where(Poi.poi_id.in_([row.poi_id for row in poi_rows]))
            )
        )
        for row in poi_rows:
            values = asdict(row)
            values["location"] = func.ST_SetSRID(
                func.ST_MakePoint(row.longitude, row.latitude),
                4326,
            )
            statement = insert(Poi).values(**values)
            statement = statement.on_conflict_do_update(
                index_elements=[Poi.poi_id],
                set_={
                    "poi_name": statement.excluded.poi_name,
                    "alias": statement.excluded.alias,
                    "address": statement.excluded.address,
                    "longitude": statement.excluded.longitude,
                    "latitude": statement.excluded.latitude,
                    "category": statement.excluded.category,
                    "source": statement.excluded.source,
                    "location": statement.excluded.location,
                    "created_at": statement.excluded.created_at,
                    "updated_at": statement.excluded.updated_at,
                },
            )
            session.execute(statement)
        session.commit()

    inserted = sum(row.poi_id not in existing_ids for row in poi_rows)
    return ImportResult(
        rows_read=len(poi_rows),
        inserted=inserted,
        updated=len(poi_rows) - inserted,
    )


def import_workbook(source_path: Path) -> ImportResult:
    return upsert_pois(read_poi_rows(source_path))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Path to Macau_Route_Database_simple.xlsx")
    args = parser.parse_args()
    source_path = args.source.resolve(strict=True)
    result = import_workbook(source_path)
    print(json.dumps(asdict(result), ensure_ascii=False))


if __name__ == "__main__":
    main()
